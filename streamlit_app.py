# streamlit_app.py — Off Axis Entertainment GFX QC (host-ready)
#
# Changes vs. the desktop build:
#  1. API key is read from st.secrets["OPENAI_API_KEY"] (set server-side on the host).
#     Falls back to a local sidebar entry only if no secret is configured (for local dev).
#  2. A shared password gate (st.secrets["APP_PASSWORD"]) protects the public URL.
#  3. The "Quit Server" button is removed (you never want a user killing the shared server).
#
# QC fixes in this revision:
#  - Reserved bug zone now uses edge-density (not raw brightness); off by default.
#  - Price-format warning lands only on rows that contain a price (no more smearing).
#  - Name-inconsistency uses word-boundary matching (no more false substring hits).
#  - Removed the redundant ALL-CAPS check (Capitalization drift fully covers it).
#  - Added "Grab frames from a video" mode (browser-side capture, J-K-L shuttle).
#
# To release an update: change APP_VERSION below, then commit and push.

import os
import json
import base64
import tempfile
import re
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from difflib import SequenceMatcher

import streamlit as st
from openai import OpenAI, RateLimitError, APIError, AuthenticationError
from PIL import Image, ImageFilter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, PatternFill, Font

# Optional dependency (dictionary spelling)
try:
    from spellchecker import SpellChecker  # pip install pyspellchecker
except Exception:
    SpellChecker = None

# -----------------------------
# App Identity / Local Storage
# -----------------------------
APP_NAME = "Off Axis Entertainment GFX QC"
APP_VERSION = "v1.5"
CONFIG_DIR = Path.home() / "Library" / "Application Support" / "OffAxisGFXQC"
CONFIG_PATH = CONFIG_DIR / "config.json"

# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title=APP_NAME, page_icon="✅", layout="wide")

# -----------------------------
# Password gate (shared team password)
# -----------------------------
def check_password() -> bool:
    """Returns True if the user is allowed in.
    If no APP_PASSWORD secret is set, the gate is disabled (open access)."""
    expected = None
    try:
        expected = st.secrets.get("APP_PASSWORD")
    except Exception:
        expected = None

    if not expected:
        # No password configured -> don't block (useful for local dev).
        return True

    if st.session_state.get("auth_ok"):
        return True

    st.title(APP_NAME)
    st.text_input("Team password", type="password", key="auth_pw")
    if st.button("Enter"):
        if st.session_state.get("auth_pw", "") == expected:
            st.session_state["auth_ok"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False

if not check_password():
    st.stop()

st.title(APP_NAME)
st.caption(APP_VERSION)

# -----------------------------
# Mode: run QC on stills, or grab frames from a video first
# -----------------------------
mode = st.radio(
    "What do you want to do?",
    ["Run QC on stills", "Grab frames from a video"],
    horizontal=True,
)

if mode == "Grab frames from a video":
    import streamlit.components.v1 as components
    st.caption(
        "Load a video, shuttle to each graphic (J/K/L), and press the Up arrow (or Capture) to grab a frame. "
        "Click Download all, then switch to 'Run QC on stills' above and upload the saved frames."
    )
    try:
        grabber_html = (Path(__file__).parent / "video_grabber.html").read_text(encoding="utf-8")
    except Exception:
        grabber_html = "<p style='color:#e6e8ec'>Could not load the video grabber.</p>"
    components.html(grabber_html, height=900, scrolling=True)
    st.stop()

# -----------------------------
# Session state
# -----------------------------
def init_state():
    st.session_state.setdefault("qc_running", False)
    st.session_state.setdefault("qc_ran", False)
    st.session_state.setdefault("qc_extracted", None)
    st.session_state.setdefault("qc_images", None)
    st.session_state.setdefault("qc_issues", None)
    st.session_state.setdefault("qc_xlsx_path", None)
    st.session_state.setdefault("qc_filename", None)
    st.session_state.setdefault("qc_sig", None)

init_state()

def reset_results():
    st.session_state.qc_running = False
    st.session_state.qc_ran = False
    st.session_state.qc_extracted = None
    st.session_state.qc_images = None
    st.session_state.qc_issues = None
    st.session_state.qc_xlsx_path = None
    st.session_state.qc_filename = None
    st.session_state.qc_sig = None

# -----------------------------
# Helpers: config load/save (local-dev fallback only)
# -----------------------------
def load_config() -> dict:
    try:
        if CONFIG_PATH.exists():
            return json.loads(CONFIG_PATH.read_text())
    except Exception:
        pass
    return {}

def save_config(cfg: dict) -> None:
    try:
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        CONFIG_PATH.write_text(json.dumps(cfg, indent=2))
    except Exception:
        # On a hosted/read-only filesystem this may fail; that's fine.
        pass

cfg = load_config()

def sanitize_filename_part(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^A-Za-z0-9_\-\.]+", "", s)
    return s[:60]

def safe_rerun():
    try:
        st.rerun()
    except Exception:
        try:
            st.experimental_rerun()
        except Exception:
            pass

def compute_run_signature(uploaded_files, series_name: str, episode_number: str) -> str:
    parts = []
    for f in (uploaded_files or []):
        try:
            parts.append(f"{f.name}:{getattr(f, 'size', '')}")
        except Exception:
            parts.append(getattr(f, "name", "file"))
    parts.sort()
    s = (series_name or "").strip()
    e = (episode_number or "").strip()
    return "|".join(parts) + f"||SERIES={s}||EP={e}"

# -----------------------------
# Resolve API key (host secret first, sidebar fallback for local dev)
# -----------------------------
def get_secret_key() -> str:
    try:
        return (st.secrets.get("OPENAI_API_KEY") or "").strip()
    except Exception:
        return ""

secret_key = get_secret_key()

# -----------------------------
# Sidebar: settings
# -----------------------------
with st.sidebar:
    st.header("Settings")

    if secret_key:
        st.subheader("OpenAI API Key")
        st.success("Key managed by host (configured server-side).")
        api_key = secret_key
    else:
        # Local-dev fallback: let the operator paste a key.
        st.subheader("OpenAI API Key")
        saved_key = cfg.get("openai_api_key", "")
        api_key_input = st.text_input(
            "API Key",
            value=saved_key,
            type="password",
            help="No host secret found; using a locally entered key (dev mode).",
        )
        colA, colB = st.columns(2)
        with colA:
            if st.button("Save Key", use_container_width=True):
                cfg["openai_api_key"] = api_key_input.strip()
                save_config(cfg)
                st.success("Saved locally.")
        with colB:
            if st.button("Clear Key", use_container_width=True):
                cfg["openai_api_key"] = ""
                save_config(cfg)
                st.warning("Cleared locally.")
        api_key = (cfg.get("openai_api_key") or api_key_input or "").strip()

    # =====================================================
    # QC CONTROLS (Spelling first)
    # =====================================================
    st.divider()
    st.subheader("QC Controls")

    st.markdown("**Spelling (FAIL)**")
    use_dictionary_spelling = st.checkbox(
        "Dictionary spelling check (FAIL if wrong)",
        value=True,
        help="Offline dictionary. Add names/brands to whitelist (and/or Known Names) to avoid false fails.",
    )
    custom_whitelist_text = st.text_area(
        "Spelling whitelist (one word per line)",
        value="Flavortown\nFieri\nTOC\nWBD\nDiscovery\nMagnolia\nFoodNetwork\n",
        height=120,
        help="Words/names/brands that should never be flagged as misspellings.",
    )

    st.markdown("**AI Checks**")
    ai_safe_check = st.checkbox("AI Title Safe (90%) + Action Safe", value=True)
    safe_be_conservative = st.checkbox(
        "Be conservative (prefer WARN over PASS)",
        value=True,
        help="If unsure, the model should mark WARN and explain briefly.",
    )

    st.markdown("**Rules Checks (Warnings)**")

    check_reserved_bug_zone = st.checkbox(
        "Reserved bug zone (warn if graphics enter lower-right area)",
        value=False,
        help="Approximate edge-density check on the lower-right network-bug area. Can trip on busy backgrounds, so it's off by default — turn on as a soft WARN.",
    )
    if check_reserved_bug_zone:
        bug_zone_width_pct = st.slider("Bug zone width (%)", 5, 30, 18)
        bug_zone_height_pct = st.slider("Bug zone height (%)", 10, 60, 40)
    else:
        bug_zone_width_pct, bug_zone_height_pct = 18, 40

    check_caps_drift = st.checkbox(
        "Capitalization drift (warn only outliers vs the batch majority)",
        value=True,
        help="Majority-based: if most frames are ALL CAPS warn the non-caps; if most are not, warn the caps.",
    )
    check_accent_drift = st.checkbox(
        "Accent/diacritic consistency (warn only outliers vs the batch majority)",
        value=True,
        help="Majority-based: warns only frames that differ from the dominant accent style.",
    )
    check_price_format = st.checkbox("Price formatting consistency ($1500 vs $1,500)", value=True)

    check_name_consistency = st.checkbox(
        "Name consistency + near-miss detection (WARN)",
        value=True,
        help="Uses Known Names to warn on variants and likely typos (double letters, swapped chars, etc.).",
    )
    known_names_text = st.text_area(
        "Known names (one per line)",
        value="Guy Fieri\nRobert Irvine\n",
        height=120,
        help="Approved spellings. QC warns on variants and close misspellings.",
    )
    name_fuzzy_threshold = st.slider(
        "Name near-miss sensitivity",
        min_value=80,
        max_value=98,
        value=92,
        help="Higher = fewer false positives. 92 is a good starting point.",
    )

    st.divider()
    st.subheader("Performance")
    max_workers = st.slider(
        "Parallel requests",
        min_value=1,
        max_value=6,
        value=3,
        help="Higher = faster, but may hit rate limits.",
    )
    image_width = st.slider("Excel image width", 150, 500, 320)
    image_height = st.slider("Excel image height", 100, 400, 190)

# -----------------------------
# Require API key
# -----------------------------
api_key = (api_key or "").strip()
if not api_key:
    st.info("No OpenAI API key available. Ask the host admin to configure OPENAI_API_KEY.")
    st.stop()

client = OpenAI(api_key=api_key)

# -----------------------------
# Project Metadata (Series / Episode)
# -----------------------------
meta_c1, meta_c2, meta_c3 = st.columns([2, 1, 2])
with meta_c1:
    series_name = st.text_input("Series Name", placeholder="e.g. Welcome to Flavortown")
with meta_c2:
    episode_number = st.text_input("Episode", placeholder="e.g. 101")
with meta_c3:
    st.caption("Series/Episode stamped into Excel + filename.")

require_meta = st.checkbox("Require Series + Episode before processing", value=False)

# -----------------------------
# Upload
# -----------------------------
uploaded_files = st.file_uploader(
    "Upload frame images (PNG/JPG).",
    type=["png", "jpg", "jpeg"],
    accept_multiple_files=True,
)

current_sig = compute_run_signature(uploaded_files, series_name, episode_number)
if st.session_state.qc_ran and st.session_state.qc_sig and current_sig != st.session_state.qc_sig:
    reset_results()

# -----------------------------
# Process controls
# -----------------------------
btn1, btn2, btn3 = st.columns([2, 1, 1])

with btn1:
    process_clicked = st.button(
        "Process Frames",
        type="primary",
        use_container_width=True,
        disabled=(not uploaded_files) or st.session_state.qc_running,
    )

with btn2:
    if st.button("Reset Results", use_container_width=True, disabled=st.session_state.qc_running):
        reset_results()
        safe_rerun()

with btn3:
    st.caption("QC Running…" if st.session_state.qc_running else "Upload → set QC → Process")

if st.session_state.qc_ran and st.session_state.qc_xlsx_path:
    st.success("QC processed. Download below. Upload new frames (or change Series/Episode) to run again.")
    try:
        with open(st.session_state.qc_xlsx_path, "rb") as f:
            st.download_button(
                "Download Excel QC Log",
                f,
                file_name=st.session_state.qc_filename or "GFX_QC.xlsx",
                use_container_width=True,
                key="download_existing",
            )
    except Exception as e:
        st.warning("Previous Excel file could not be opened. Upload frames again and re-run.")
        st.code(str(e))

if not process_clicked:
    st.stop()

if require_meta and ((not (series_name or "").strip()) or (not (episode_number or "").strip())):
    st.warning("Please enter **Series Name** and **Episode** (or uncheck 'Require Series + Episode').")
    st.stop()

st.session_state.qc_running = True

# -----------------------------
# Vision extract (+ Safe)
# -----------------------------
def encode_image(uploaded_file) -> str:
    uploaded_file.seek(0)
    return base64.b64encode(uploaded_file.read()).decode("utf-8")

def vision_extract(image_b64: str) -> dict:
    safe_rules = ""
    if ai_safe_check:
        safe_rules = (
            "SAFE AREA CHECKS:\n"
            "TITLE SAFE RULE (90%):\n"
            "- Title-safe area is the INNER 90% of the image (width and height).\n"
            "- That means a 5% margin on ALL sides is NOT title-safe.\n"
            "- If any essential text/logo/critical info touches or enters that 5% border, mark title_safe WARN or FAIL.\n"
            "- WARN if close/uncertain; FAIL if clearly in unsafe margin or at risk of clipping.\n"
            "- Unsafe border zones: left<5%, right>95%, top<5%, bottom>95%.\n\n"
            "ACTION SAFE:\n"
            "- Use conservative action-safe guidance (flag important action/logo too close to edges).\n"
        )

    resp = client.chat.completions.create(
        model="gpt-4.1-mini",
        messages=[
            {
                "role": "system",
                "content": (
                    "Return ONLY valid JSON.\n\n"
                    "Required keys:\n"
                    "- timecode: string\n"
                    "- gfx_text: string (preserve line breaks exactly)\n"
                    "- title_safe: PASS|WARN|FAIL\n"
                    "- action_safe: PASS|WARN|FAIL\n"
                    "- safe_notes: string\n\n"
                    "gfx_text must preserve line breaks exactly as seen.\n\n"
                    + safe_rules
                ),
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            f"Extract timecode + GFX text. "
                            f"Safe-check={ai_safe_check} (title safe 90%). conservative={safe_be_conservative}. "
                            f"Output JSON only."
                        ),
                    },
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}},
                ],
            },
        ],
        response_format={"type": "json_object"},
    )

    data = json.loads(resp.choices[0].message.content)

    timecode = (data.get("timecode") or "").strip() or "Not legible"
    gfx_text = (data.get("gfx_text") or "").strip() or "Not legible"
    if gfx_text.startswith('"') and gfx_text.endswith('"'):
        gfx_text = gfx_text[1:-1].strip()

    title_safe = (data.get("title_safe") or "PASS").strip().upper()
    action_safe = (data.get("action_safe") or "PASS").strip().upper()
    safe_notes = (data.get("safe_notes") or "").strip()

    if title_safe not in ("PASS", "WARN", "FAIL"):
        title_safe = "PASS"
    if action_safe not in ("PASS", "WARN", "FAIL"):
        action_safe = "PASS"

    if ai_safe_check and safe_be_conservative and safe_notes:
        if title_safe == "PASS" and "uncertain" in safe_notes.lower():
            title_safe = "WARN"
        if action_safe == "PASS" and "uncertain" in safe_notes.lower():
            action_safe = "WARN"

    return {
        "timecode": timecode,
        "gfx_text": gfx_text,
        "title_safe": title_safe,
        "action_safe": action_safe,
        "safe_notes": safe_notes,
    }

# -----------------------------
# Reserved bug zone detector
# -----------------------------
def detect_reserved_bug_zone_violation(img: Image.Image, width_pct: int = 18, height_pct: int = 40, edge_threshold: float = 18.0) -> bool:
    """Approximate: flag if the reserved lower-right zone contains dense, sharp
    edges (text / graphic elements) instead of smooth background video.
    Edge density responds to graphics and text, and (unlike the old brightness
    test) ignores a merely bright but smooth background. Still a heuristic —
    busy live backgrounds can trip it — so it's meant as a soft WARN and is
    off by default."""
    w, h = img.size
    x0 = int(w * (1 - (width_pct / 100.0)))
    y0 = int(h * (1 - (height_pct / 100.0)))
    region = img.crop((x0, y0, w, h)).convert("L")
    edges = region.filter(ImageFilter.FIND_EDGES)
    pixels = list(edges.getdata())
    if not pixels:
        return False
    mean_edge = sum(pixels) / len(pixels)
    return mean_edge > edge_threshold

# -----------------------------
# ALL CAPS helpers
# -----------------------------
def upper_ratio(t: str) -> float:
    letters = [c for c in t if c.isalpha()]
    if not letters:
        return 0.0
    return sum(1 for c in letters if c.isupper()) / len(letters)

def is_all_caps_text(t: str) -> bool:
    letters = [c for c in t if c.isalpha()]
    if len(letters) < 4:
        return False
    return upper_ratio(t) >= 0.90

# -----------------------------
# Dictionary spelling helpers
# -----------------------------
def tokenize_for_spellcheck(text: str) -> list[str]:
    return re.findall(r"[A-Za-z]+(?:'[A-Za-z]+)?", text or "")

def build_whitelist(custom_whitelist: str, known_names: str) -> set[str]:
    wl = set()
    for line in (custom_whitelist or "").splitlines():
        w = line.strip()
        if w:
            wl.add(w.lower())
    for line in (known_names or "").splitlines():
        line = line.strip()
        if not line:
            continue
        for tok in re.findall(r"[A-Za-z]+(?:'[A-Za-z]+)?", line):
            wl.add(tok.lower())
    return wl

def dictionary_spellcheck_fail(text: str, whitelist: set[str], spell) -> tuple[bool, str]:
    tokens = tokenize_for_spellcheck(text)
    cleaned = []
    for t in tokens:
        tl = t.lower()
        if len(tl) <= 2:
            continue
        if tl in whitelist:
            continue
        if t.isupper() and len(t) <= 5:
            continue
        cleaned.append(t)

    if not cleaned:
        return False, ""

    misspelled = spell.unknown([t.lower() for t in cleaned])
    if not misspelled:
        return False, ""

    parts = []
    for w in sorted(misspelled):
        if w in whitelist:
            continue
        sug = spell.correction(w) or ""
        parts.append(f"{w}→{sug}" if sug and sug != w else w)

    parts = parts[:8]
    return True, "Misspellings: " + ", ".join(parts)

# -----------------------------
# Name consistency helpers
# -----------------------------
def normalize_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("’", "'")
    return s

def name_signature(s: str) -> str:
    return re.sub(r"[^a-z]", "", (s or "").lower())

def similarity(a: str, b: str) -> int:
    return int(round(100 * SequenceMatcher(None, a, b).ratio()))

def build_known_names(known_names_text_in: str) -> list[dict]:
    names = []
    for line in (known_names_text_in or "").splitlines():
        canon = normalize_name(line)
        if canon:
            names.append({"canon": canon, "sig": name_signature(canon)})
    return names

def extract_name_candidates(text: str) -> list[str]:
    t = normalize_name(text)
    cands = re.findall(r"\b[A-Z][A-Za-z'’\-\.]+(?:\s+[A-Z][A-Za-z'’\-\.]+){1,3}\b", t)
    return [normalize_name(c) for c in cands]

# -----------------------------
# Accent helper
# -----------------------------
def has_non_ascii(s: str) -> bool:
    try:
        s.encode("ascii")
        return False
    except Exception:
        return True

# -----------------------------
# Issues (single column)
# -----------------------------
def detect_issues(extracted_texts: list, pil_images: list) -> list:
    issues = [""] * len(extracted_texts)
    texts = [e.get("gfx_text", "") for e in extracted_texts]

    whitelist = build_whitelist(custom_whitelist_text, known_names_text)
    spell = None
    if use_dictionary_spelling:
        if SpellChecker is None:
            st.warning("Dictionary spelling enabled but 'pyspellchecker' not installed (pip install pyspellchecker).")
        else:
            spell = SpellChecker()

    if check_name_consistency:
        known = build_known_names(known_names_text)
        variants_by_canon = {k["canon"]: set() for k in known}
        near_miss_notes = [[] for _ in range(len(texts))]

        for i, t in enumerate(texts):
            candidates = extract_name_candidates(t)
            for cand in candidates:
                cand_sig = name_signature(cand)

                matched_exact = False
                for k in known:
                    if cand_sig == k["sig"]:
                        variants_by_canon[k["canon"]].add(cand)
                        matched_exact = True
                        break

                if (not matched_exact) and known:
                    best = None
                    best_score = 0
                    for k in known:
                        sc = similarity(cand_sig, k["sig"])
                        if sc > best_score:
                            best_score = sc
                            best = k
                    if best and best_score >= name_fuzzy_threshold and cand_sig != best["sig"]:
                        near_miss_notes[i].append(f"{cand}≈{best['canon']} ({best_score}%)")

        def contains_variant(text_in: str, variant: str) -> bool:
            return re.search(r"\b" + re.escape(variant) + r"\b", text_in) is not None

        for canon, vars_seen in variants_by_canon.items():
            if len(vars_seen) > 1:
                pretty = ", ".join(sorted(vars_seen))
                for i, t in enumerate(texts):
                    if any(contains_variant(t, v) for v in vars_seen):
                        issues[i] += f"WARN — Name inconsistency ({canon}: {pretty}) | "

        for i, notes in enumerate(near_miss_notes):
            if notes:
                issues[i] += "WARN — Possible name misspelling (" + "; ".join(notes[:3]) + ") | "

    for i, e in enumerate(extracted_texts):
        if ai_safe_check:
            ts = (e.get("title_safe") or "PASS").upper()
            ac = (e.get("action_safe") or "PASS").upper()
            if ts in ("WARN", "FAIL"):
                issues[i] += f"{ts} — Title Safe (90%) | "
            if ac in ("WARN", "FAIL"):
                issues[i] += f"{ac} — Action Safe | "

        if check_reserved_bug_zone and pil_images and pil_images[i] is not None:
            if detect_reserved_bug_zone_violation(pil_images[i], bug_zone_width_pct, bug_zone_height_pct):
                issues[i] += "WARN — Graphics entered reserved bug zone | "

        if use_dictionary_spelling and spell is not None:
            has_err, note = dictionary_spellcheck_fail(e.get("gfx_text", ""), whitelist, spell)
            if has_err:
                issues[i] += f"FAIL — Spelling Error ({note}) | "

    if check_caps_drift and len(texts) >= 2:
        caps_flags = []
        for t in texts:
            letters = [c for c in t if c.isalpha()]
            if len(letters) < 4:
                caps_flags.append(None)
            else:
                caps_flags.append(is_all_caps_text(t))

        voted = [v for v in caps_flags if v is not None]
        if len(voted) >= 2:
            caps_true = sum(1 for v in voted if v)
            caps_false = sum(1 for v in voted if not v)
            total = caps_true + caps_false

            majority_ratio = max(caps_true, caps_false) / total if total else 0.0
            if majority_ratio >= 0.70 and caps_true > 0 and caps_false > 0:
                expected = (caps_true >= caps_false)
                for i, v in enumerate(caps_flags):
                    if v is None:
                        continue
                    if v != expected:
                        issues[i] += "WARN — Capitalization drift | "

    if check_accent_drift and len(texts) >= 2:
        accent_flags = []
        for t in texts:
            letters = [c for c in t if c.isalpha()]
            if len(letters) < 4:
                accent_flags.append(None)
            else:
                accent_flags.append(has_non_ascii(t))

        voted = [v for v in accent_flags if v is not None]
        if len(voted) >= 2:
            acc_true = sum(1 for v in voted if v)
            acc_false = sum(1 for v in voted if not v)
            total = acc_true + acc_false

            majority_ratio = max(acc_true, acc_false) / total if total else 0.0
            if majority_ratio >= 0.70 and acc_true > 0 and acc_false > 0:
                expected = (acc_true >= acc_false)
                for i, v in enumerate(accent_flags):
                    if v is None:
                        continue
                    if v != expected:
                        issues[i] += "WARN — Accent/diacritic consistency | "

    if check_price_format:
        per_row_prices = []
        all_prices = []
        for t in texts:
            found = re.findall(r"\$\s?\d[\d,]*(?:\.\d{2})?", t)
            per_row_prices.append(found)
            all_prices += found
        if all_prices and any("," in p for p in all_prices) and any("," not in p for p in all_prices):
            for i, found in enumerate(per_row_prices):
                if found:
                    issues[i] += "WARN — Price formatting inconsistency | "

    issues = [x[:-3] if x.endswith(" | ") else (x or "") for x in issues]
    return [x if x else "None" for x in issues]

# -----------------------------
# Excel (4 columns + series/ep at top)
# -----------------------------
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
META_LABEL_FONT = Font(bold=True)

def write_excel_minimal(rows: list, issue_strings: list, images: list, series: str, episode: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Log"

    series = (series or "").strip()
    episode = (episode or "").strip()

    ws["A1"] = "Series"
    ws["B1"] = series or "—"
    ws["A2"] = "Episode"
    ws["B2"] = episode or "—"
    ws["A1"].font = META_LABEL_FONT
    ws["A2"].font = META_LABEL_FONT
    ws["A3"] = ""

    ws["A4"] = "Timecode"
    ws["B4"] = "On-Screen Text"
    ws["C4"] = "Issue"
    ws["D4"] = "Image"
    for cell in ("A4", "B4", "C4", "D4"):
        ws[cell].font = Font(bold=True)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 46

    start_row = 5
    tmp_paths = []
    for idx, r in enumerate(rows, start=start_row):
        src_i = idx - start_row
        ws[f"A{idx}"] = r.get("timecode", "")
        ws[f"B{idx}"] = r.get("gfx_text", "")
        issue_text = issue_strings[src_i] if issue_strings else "None"
        ws[f"C{idx}"] = issue_text

        ws[f"B{idx}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws[f"C{idx}"].alignment = Alignment(wrap_text=True, vertical="center")

        if issue_text and issue_text != "None":
            ws[f"C{idx}"].fill = FAIL_FILL if "FAIL" in issue_text.upper() else WARN_FILL

        try:
            tmp_img = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            images[src_i].save(tmp_img.name)
            tmp_paths.append(tmp_img.name)
            xl_img = XLImage(tmp_img.name)
            xl_img.width = int(image_width)
            xl_img.height = int(image_height)
            ws.add_image(xl_img, f"D{idx}")
        except Exception:
            ws[f"D{idx}"] = "(image failed)"
        ws.row_dimensions[idx].height = max(120, int(image_height * 0.75))

    out_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(out_xlsx.name)
    return out_xlsx.name

# -----------------------------
# Processing run (Process clicked)
# -----------------------------
tabs = st.tabs(["QC Log", "Summary"])
with tabs[0]:
    st.write(f"Frames uploaded: **{len(uploaded_files)}**")
    st.write(
        f"Series: **{(series_name or '').strip() or '—'}**  |  "
        f"Episode: **{(episode_number or '').strip() or '—'}**"
    )

    progress = st.progress(0)
    status = st.empty()

    extracted = [None] * len(uploaded_files)
    pil_images = [None] * len(uploaded_files)

    def worker(i, f):
        img_b64 = encode_image(f)
        f.seek(0)
        pil = Image.open(f).convert("RGB")
        data = vision_extract(img_b64)
        return i, data, pil

    try:
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = [ex.submit(worker, i, f) for i, f in enumerate(uploaded_files)]
            done = 0
            total = len(futures)
            for fut in as_completed(futures):
                i, data, pil = fut.result()
                extracted[i] = data
                pil_images[i] = pil
                done += 1
                progress.progress(done / total)
                status.write(f"Processing {done}/{total}...")

    except AuthenticationError:
        st.session_state.qc_running = False
        st.error("OpenAI API key is invalid. Ask the host admin to update OPENAI_API_KEY.")
        st.stop()
    except RateLimitError as e:
        st.session_state.qc_running = False
        st.error("OpenAI quota/rate limit hit. Try fewer parallel requests or check billing/quota.")
        st.code(str(e))
        st.stop()
    except APIError as e:
        st.session_state.qc_running = False
        st.error("OpenAI API error. Try again in a moment.")
        st.code(str(e))
        st.stop()
    except Exception as e:
        st.session_state.qc_running = False
        st.error("Unexpected error while processing frames.")
        st.code(str(e))
        st.stop()

    status.write("Running QC checks...")
    per_row_issues = detect_issues(extracted, pil_images)

    status.write("Building Excel...")
    xlsx_path = write_excel_minimal(
        extracted,
        per_row_issues,
        pil_images,
        series=(series_name or "").strip(),
        episode=(episode_number or "").strip(),
    )

    safe_series = sanitize_filename_part(series_name) or "Series"
    safe_ep = sanitize_filename_part(episode_number) or "EP"
    out_name = f"{safe_series}_EP{safe_ep}_GFX_QC.xlsx"

    st.session_state.qc_ran = True
    st.session_state.qc_running = False
    st.session_state.qc_extracted = extracted
    st.session_state.qc_images = pil_images
    st.session_state.qc_issues = per_row_issues
    st.session_state.qc_xlsx_path = xlsx_path
    st.session_state.qc_filename = out_name
    st.session_state.qc_sig = current_sig

    st.success("QC complete. Refreshing…")
    safe_rerun()
    st.stop()

with tabs[1]:
    st.info("Summary appears after processing completes.")
